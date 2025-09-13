import openpyxl
import json
import time
import random
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from fake_useragent import UserAgent
import pypartpicker

# ======== ScraperAPI Proxy 設定 ========
api_key = "66edc8a76a69bb88bf657e76121eed25"
proxy_url = f"http://proxy.scraperapi.com:8001?api_key={api_key}"

# ======== 讀取商品連結 Excel ========
filename = "pcpartpicker_links_all.xlsx"
wb = openpyxl.load_workbook(filename)
all_product_links = {}

for sheet in wb.sheetnames:
    ws = wb[sheet]
    links = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    all_product_links[sheet] = links

# ======== 建立 pypartpicker client ========
client = pypartpicker.Client()

# ======== Selenium 設定 ========
ua = UserAgent()
user_agent = ua.random
options = Options()
# options.add_argument("--headless")
options.add_argument(f"user-agent={user_agent}")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--window-size=1920,1080")
options.add_argument(f'--proxy-server={proxy_url}')
driver = webdriver.Chrome(options=options)

# ======== 主爬蟲邏輯 ========
try:
    for category, links in all_product_links.items():
        print(f"開始處理分類：{category}")
        data_list = []

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.append(["商品名稱", "商品價格", "商品規格", "商品評分", "商品連結", "買家評論"])

        for link in links:
            print(f"進入商品連結：{link}")
            driver.get(link)
            time.sleep(random.uniform(3, 5))

            try:
                # 商品名稱
                name = driver.find_element(By.CLASS_NAME, "pageTitle").text

                # 商品價格
                try:
                    result = client.get_part(link)
                    price = result.cheapest_price.total if result.cheapest_price else "N/A"
                except Exception as e:
                    print(f"價格取得失敗：{e}")
                    price = "N/A"

                # 商品規格（key-value）
                specs = {}
                spec_blocks = driver.find_elements(By.CSS_SELECTOR, "div.block.xs-hide.md-block.specs div.group.group--spec")
                for block in spec_blocks:
                    try:
                        title = block.find_element(By.CLASS_NAME, "group__title").text.strip()
                        content = block.find_element(By.CLASS_NAME, "group__content").text.strip()
                        specs[title] = content
                    except:
                        continue

                # 商品評分
                try:
                    rating_ul = driver.find_element(By.CSS_SELECTOR, "div.actionBox__ratings ul.product--rating")
                    lis = rating_ul.find_elements(By.TAG_NAME, "li")
                    rating = lis[-1].text.strip() if lis else "N/A"
                except:
                    rating = "N/A"

                # 買家評論
                try:
                    reviews = driver.find_elements(By.CSS_SELECTOR, "div.partReviews__review div.partReviews__writeup.markdown")
                    comments = [r.text.strip() for r in reviews if r.text.strip()]
                except:
                    comments = []

                product_data = {
                    "name": name,
                    "price": price,
                    "specs": specs,
                    "rating": rating,
                    "link": link,
                    "comments": comments,
                }

                data_list.append(product_data)

                # 存入 Excel
                ws_out.append([
                    name,
                    price,
                    json.dumps(specs, ensure_ascii=False),
                    rating,
                    link,
                    "\n\n".join(comments)
                ])

                print(f"成功抓取：{product_data}")
            except:
                print
            
            time.sleep(random.uniform(5, 7))

        # 儲存 JSON
        json_out_name = f"data/{category}.json"
        with open(json_out_name, "w", encoding="utf-8") as f:
            json.dump(data_list, f, indent=2, ensure_ascii=False)

finally:
    driver.quit()
